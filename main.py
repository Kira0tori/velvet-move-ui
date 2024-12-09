import streamlit as st
import streamlit.components.v1 as components
import itertools
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import matplotlib.animation as animation
from scipy.signal import find_peaks
from angle_calc import calculate_signed_angle_between_vectors


def main():
    st.image("VIT_2024_square_white.png")
    #Add title and text explenation
    st.title("Post-traitement des données Velvet-Move")
    file = st.file_uploader("Please upload your exam file below (.xlsx format)")
    if file is not None:
        wb = openpyxl.load_workbook(filename=file, data_only=True)
        ws = wb.active
        px = 1 / plt.rcParams['figure.dpi']  # pixel in inches
        fig = plt.figure()  # create the canvas for upcoming plots
        sensor_list = pre_process_xlsx(ws)
        i = 1
        sensor1 = process_xlsx(ws, str(sensor_list[0]))
        sensor2 = process_xlsx(ws, str(sensor_list[1]))
        joint_1, joint_2 = coord_extrapolate(sensor1, sensor2)
        angles, sensor1_dist, sensor2_dist = angle_analysis(joint_1, joint_2)
        #angle_ax = fig.add_subplot(2, 1, 1)
        np_angles = np.asarray(angles)
        peaks, _ = find_peaks(np_angles, width=5)
       #for element in peaks:
            #angle_ax.annotate('%3.1f °' % np_angles[element], xy=(element, np_angles[element]), xytext=(10,0),
                              #textcoords='offset points')
        # Adding annotation
        #angle_ax.plot(np_angles)
        #angle_ax.plot(peaks, np_angles[peaks], "x")
        st.header("Evolution de l'angle au cours de l'exercice (°)",divider = True)
        st.line_chart(angles)
        col1, col2 = st.columns(2)
        col1.write("Angles maximum atteints lors de l'exercice")
        col1.write(np_angles[peaks])
        col2.write("Angle moyen lors de l'exercice")
        col2.write("%3.1f" % np.mean(np_angles[peaks])+" °")
        st.header("Mouvement de l'articulation au cours de l'exercice",divider = True)

        plot_joint(joint_1, joint_2, fig, angles)


def pre_process_xlsx(ws): #return a list of all different sensors in the file
    # "data_only = True" force openpyxl to read values and not formulas
    sensor_list = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=5):
        if row[0].value not in sensor_list:
            sensor_list.append(row[0].value)
        else:
            break
    return sensor_list


def process_xlsx(ws, point_name):
    data_name = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=5):
        if row[0].value == point_name:
            p = (row[1].value, row[2].value, row[3].value)
            data_name.append(p)
    return data_name


def angle_analysis(sensor1, sensor2):
    angles = []
    sensor1_dist = []
    sensor2_dist = []
    origin = (0, 0, 0)
    for i in range(len(sensor2)):
        start1 = origin
        end1 = sensor1[i]
        start2 = end1
        end2 = sensor2[i]
        angle = calculate_signed_angle_between_vectors(start1, end1, start2, end2)
        angles.append(angle)
        sensor1_dist.append(np.linalg.norm(np.asarray(end1) - np.asarray(start1)))
        sensor2_dist.append(np.linalg.norm(np.asarray(end2) - np.asarray(start2)))

    return angles, sensor1_dist, sensor2_dist


def coord_extrapolate(sensor1, sensor2):
    joint_1 = [(coord[0] * 2, coord[1] * 2, coord[2] * 2) for coord in sensor1]
    joint_2 = []
    for i, coord in enumerate(sensor2):
        new_x = coord[0] + (coord[0] - joint_1[i][0])
        new_y = coord[1] + (coord[1] - joint_1[i][1])
        new_z = coord[2] + (coord[2] - joint_1[i][2])
        joint_2.append((new_x, new_y, new_z))
    return joint_1, joint_2


def plot_joint(sensor1, sensor2, fig, angles):
    min_dim, max_dim = axis_calc((sensor1, sensor2))

    ax = fig.add_subplot(1, 1, 1, projection='3d')
    ax.set_title("Joint Plot")
    ax.set_proj_type('ortho')
    ax.set_xlabel('X')
    ax.set_ylabel('Y')
    ax.set_zlabel('Z')
    ax.set_xlim([min_dim, max_dim])
    ax.set_ylim([min_dim, max_dim])
    ax.set_zlim([min_dim, 0])
    ax.set_box_aspect([1, 1, 1])
    label = ax.text2D(1, 1, "2D Text", transform=ax.transAxes)

    line, = ax.plot([], [], [])

    def init():
        line.set_data([], [])
        line.set_3d_properties([])
        return line,

    def animate(i):
        x_values = [0, sensor1[i][0], sensor2[i][0]]
        y_values = [0, sensor1[i][1], sensor2[i][1]]
        z_values = [0, sensor1[i][2], sensor2[i][2]]
        line.set_data(x_values, z_values)
        line.set_3d_properties(y_values)

        label.set_text("Joint Angle : \n %5.1f" % angles[i])  # update the label at each frame
        return line,

    # define the animation on the fig via the animate function, each frame called at interval(in ms)
    # init_func is called first and played on repeat
    ani = animation.FuncAnimation(fig, animate, interval=50, frames=len(sensor2),
                                  init_func=init, repeat=True)
    components.html(ani.to_jshtml(), height=2000)
    #st.pyplot(fig)
    return ax


def axis_calc(sensors):
    all_data = list(itertools.chain.from_iterable(sensors[0] + sensors[1]))
    min_dim = min(all_data)
    max_dim = max(all_data)
    return min_dim, max_dim


if __name__ == '__main__':
    main()
